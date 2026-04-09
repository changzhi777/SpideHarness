# Copyright (C) 2026 IoTchange - All Rights Reserved
# Author: 外星动物（常智） / IoTchange / 14455975@qq.com
"""多格式数据导出器 — CSV / JSON / JSONL / Excel.

用法:
    from spide.storage.exporter import DataExporter

    exporter = DataExporter(output_dir="data/export")
    await exporter.export_json(topics, filename="weibo_hot")
    await exporter.export_csv(topics, filename="weibo_hot")
    await exporter.export_excel(topics, filename="weibo_hot")
    await exporter.export_jsonl(topics, filename="weibo_hot")
"""

from __future__ import annotations

import asyncio
import csv
import json
import time
from pathlib import Path
from typing import Any

from pydantic import BaseModel

from spide.exceptions import StorageError
from spide.logging import get_logger

logger = get_logger(__name__)


class DataExporter:
    """多格式数据导出器."""

    def __init__(self, *, output_dir: str = "data/export") -> None:
        self._output_dir = Path(output_dir)

    def _ensure_dir(self) -> None:
        """确保输出目录存在."""
        self._output_dir.mkdir(parents=True, exist_ok=True)

    @staticmethod
    def _model_to_dict(item: BaseModel) -> dict[str, Any]:
        """Pydantic 模型转可序列化字典."""
        return item.model_dump(mode="json")

    async def export_json(
        self,
        items: list[BaseModel],
        *,
        filename: str = "export",
        indent: int = 2,
    ) -> Path:
        """导出为 JSON 文件."""
        self._ensure_dir()
        filepath = self._output_dir / f"{filename}.json"

        data = [self._model_to_dict(item) for item in items]
        await asyncio.to_thread(self._write_json, filepath, data, indent)

        logger.debug("exported_json", path=str(filepath), count=len(items))
        return filepath

    async def export_jsonl(
        self,
        items: list[BaseModel],
        *,
        filename: str = "export",
    ) -> Path:
        """导出为 JSONL（每行一个 JSON 对象）."""
        self._ensure_dir()
        filepath = self._output_dir / f"{filename}.jsonl"

        lines = [json.dumps(self._model_to_dict(item), ensure_ascii=False) for item in items]
        await asyncio.to_thread(self._write_text, filepath, "\n".join(lines) + "\n")

        logger.debug("exported_jsonl", path=str(filepath), count=len(items))
        return filepath

    async def export_csv(
        self,
        items: list[BaseModel],
        *,
        filename: str = "export",
    ) -> Path:
        """导出为 CSV 文件."""
        if not items:
            raise StorageError("导出数据为空")

        self._ensure_dir()
        filepath = self._output_dir / f"{filename}.csv"

        data = [self._model_to_dict(item) for item in items]
        await asyncio.to_thread(self._write_csv, filepath, data)

        logger.debug("exported_csv", path=str(filepath), count=len(items))
        return filepath

    async def export_excel(
        self,
        items: list[BaseModel],
        *,
        filename: str = "export",
        sheet_name: str = "Sheet1",
    ) -> Path:
        """导出为 Excel (.xlsx) 文件."""
        if not items:
            raise StorageError("导出数据为空")

        self._ensure_dir()
        filepath = self._output_dir / f"{filename}.xlsx"

        start = time.monotonic()
        data = [self._model_to_dict(item) for item in items]
        await asyncio.to_thread(self._write_excel, filepath, data, sheet_name)
        duration_ms = (time.monotonic() - start) * 1000
        logger.debug("export_excel_duration", duration_ms=round(duration_ms, 1), record_count=len(items), path=str(filepath))

        logger.debug("exported_excel", path=str(filepath), count=len(items))
        return filepath

    async def export(
        self,
        items: list[BaseModel],
        *,
        filename: str = "export",
        fmt: str = "json",
    ) -> Path:
        """按指定格式导出.

        Args:
            items: 数据列表
            filename: 文件名（不含扩展名）
            fmt: 导出格式 (json/jsonl/csv/excel)

        Returns:
            导出文件路径
        """
        exporters = {
            "json": self.export_json,
            "jsonl": self.export_jsonl,
            "csv": self.export_csv,
            "excel": self.export_excel,
            "xlsx": self.export_excel,
        }
        exporter = exporters.get(fmt)
        if not exporter:
            raise StorageError(f"不支持的导出格式: {fmt}，可选: {', '.join(exporters.keys())}")

        return await exporter(items, filename=filename)  # type: ignore[operator]

    # -----------------------------------------------------------------------
    # 同步写入方法（通过 asyncio.to_thread 调用）
    # -----------------------------------------------------------------------

    @staticmethod
    def _write_json(filepath: Path, data: list[dict], indent: int) -> None:
        with open(filepath, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=indent)

    @staticmethod
    def _write_text(filepath: Path, text: str) -> None:
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(text)

    @staticmethod
    def _write_csv(filepath: Path, data: list[dict]) -> None:
        if not data:
            return
        fieldnames = list(data[0].keys())
        with open(filepath, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            for row in data:
                clean_row = {}
                for k, v in row.items():
                    if v is None:
                        clean_row[k] = ""
                    elif isinstance(v, (list, dict)):
                        clean_row[k] = json.dumps(v, ensure_ascii=False)
                    else:
                        clean_row[k] = v
                writer.writerow(clean_row)

    @staticmethod
    def _write_excel(filepath: Path, data: list[dict], sheet_name: str) -> None:
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        if not data:
            wb.save(filepath)
            return

        # 表头
        headers = list(data[0].keys())
        ws.append(headers)

        # 数据行
        for row_data in data:
            row = []
            for h in headers:
                val = row_data.get(h)
                if isinstance(val, (list, dict)):
                    val = json.dumps(val, ensure_ascii=False)
                elif val is None:
                    val = ""
                row.append(val)
            ws.append(row)

        # 自动调整列宽
        for col_idx, header in enumerate(headers, 1):
            max_len = len(str(header))
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        max_len = max(max_len, min(len(str(cell.value)), 50))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_len + 2

        wb.save(filepath)

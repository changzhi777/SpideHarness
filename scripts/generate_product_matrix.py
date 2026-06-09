#!/usr/bin/env python3
"""生成 SpideHarness 产品矩阵 Excel 表"""

import openpyxl
from openpyxl.styles import (
    Font, Fill, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "产品矩阵"

# ─── 配色方案 ────────────────────────────────────────
COL_HEADER_BG   = "1E3A5F"   # 深蓝 header
COL_HEADER_FG   = "FFFFFF"   # 白色文字
COL_ALT_ROW     = "E8F0F7"   # 浅蓝交替行
COL_BORDER      = "B8CCE4"   # 边框蓝
COL_STATUS_DONE = "27AE60"   # 绿-完成
COL_STATUS_DEV  = "F39C12"   # 橙-dev
COL_STATUS_ALPHA= "9B59B6"   # 紫-内测

# ─── 样式工厂 ────────────────────────────────────────
def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def make_border(color=COL_BORDER):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def make_header_cell(ws, row, col, value):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name="Microsoft YaHei", bold=True, color=COL_HEADER_FG, size=11)
    cell.fill = make_fill(COL_HEADER_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = make_border("FFFFFF")
    return cell

def make_data_cell(ws, row, col, value, bold=False, align="left", bg=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name="Microsoft YaHei", bold=bold, size=10)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border = make_border()
    if bg:
        cell.fill = make_fill(bg)
    elif row % 2 == 0:
        cell.fill = make_fill(COL_ALT_ROW)
    return cell

# ─── 数据定义 ────────────────────────────────────────
headers = [
    "功能模块", "产品定位", "应用场景", "市场需求",
    "开发程度", "完成度", "商业模式"
]

# 功能完整数据
data = [
    # (功能, 产品定位, 应用场景, 市场需求, 开发程度, 完成度, 商业模式)
    (
        "热搜采集\n(微博/百度/抖音/知乎/B站)",
        "实时热点追踪\n5平台聚合",
        "自媒体追热点\n品牌舆情监控",
        "热点发现是刚需\n用户粘性入口",
        "上线",
        "95%",
        "开源免费版\n核心功能"
    ),
    (
        "深度采集\n(小红书/抖音/快手/B站等7平台)",
        "内容原生态获取\nUGC深度分析",
        "竞品分析\n行业研究",
        "内容创作素材\nKOL营销洞察",
        "dev可运行",
        "85%",
        "开源免费版\n核心功能"
    ),
    (
        "AI智能分析\n(摘要/情感/趋势/策略)",
        "AI驱动的\n信息增值",
        "内容策划\n舆情研判",
        "降本增效\n差异化竞争",
        "上线",
        "90%",
        "开源免费版\n核心功能"
    ),
    (
        "词云生成\n(jieba分词+wordcloud)",
        "可视化呈现\n关键词洞察",
        "报告展示\n趋势呈现",
        "汇报场景\n展示需求",
        "上线",
        "95%",
        "开源免费版"
    ),
    (
        "定时任务调度\n(Cron-like)",
        "自动化运营\n解放人力",
        "7x24监控\n定时报告",
        "企业级刚需\n运营自动化",
        "上线",
        "90%",
        "开源免费版"
    ),
    (
        "数据导出\n(JSON/CSV/Excel)",
        "多格式兼容\n数据打通",
        "二次分析\n报告制作",
        "数据资产化\n报表需求",
        "上线",
        "95%",
        "开源免费版"
    ),
    (
        "MCP协议\n(Server/Client)",
        "AI原生架构\n模型即服务",
        "AI助手集成\nClaude/GPT接入",
        "开发者生态\nMCP是标配",
        "上线",
        "85%",
        "开源免费版"
    ),
    (
        "MQTT通讯\n(EMQX Cloud TLS)",
        "消息实时推送\n事件驱动",
        "系统集成\nWebhook通知",
        "企业集成\nIoT场景",
        "上线",
        "90%",
        "开源免费版"
    ),
    (
        "浏览器自动化\n(OpenCLI 79+适配器)",
        "无API场景覆盖\n登录态复用",
        "复杂交互采集\n社交媒体抓取",
        "数据采集完整性\n差异化能力",
        "dev可运行",
        "80%",
        "开源免费版"
    ),
    (
        "智能搜索路由\n(AI+60+网站多源)",
        "信息发现引擎\n精准检索",
        "深度调研\n竞品分析",
        "信息差竞争力\n效率提升",
        "dev可运行",
        "75%",
        "展屏版/团队版"
    ),
    (
        "适配器生态\n(Explorer/Oneshot/Autofix)",
        "开发者友好\n零成本扩展",
        "新平台适配\n定制采集",
        "生态护城河\n长期竞争力",
        "正在开发",
        "60%",
        "展屏版/团队版"
    ),
    (
        "大数据可视化展屏\n(React+ECharts)",
        "B端核心产品\n差异化体验",
        "展厅展示\n数据大屏",
        "企业采购决策\n高溢价能力",
        "正在开发",
        "20%",
        "展屏版¥99/月"
    ),
    (
        "关键词实时告警\n(多通道通知)",
        "主动运营\n危机响应",
        "舆情监控\n危机公关",
        "企业刚需\n付费意愿强",
        "正在开发",
        "15%",
        "团队版¥299/月"
    ),
    (
        "自动日报/周报\n(AI生成推送)",
        "运营自动化\n效率提升",
        "日常汇报\n管理决策",
        "老板刚需\n降本增效",
        "正在开发",
        "10%",
        "团队版¥299/月"
    ),
    (
        "Web Dashboard\n(Vue3+FastAPI)",
        "团队协作平台\n用户管理",
        "团队分工\n权限管理",
        "团队版核心\n企业级需求",
        "正在开发",
        "5%",
        "团队版¥299/月"
    ),
    (
        "历史趋势对比\n(多维度分析)",
        "深度BI能力\n决策支持",
        "策略复盘\n效果评估",
        "分析深度\n差异化价值",
        "正在开发",
        "5%",
        "企业版¥2999/月"
    ),
    (
        "情感指数API\n(舆情量化指标)",
        "数据产品化\nAPI经济",
        "金融舆情\n量化策略",
        "B端API收入\n高毛利",
        "规划中",
        "0%",
        "企业版API"
    ),
    (
        "热点→概念股关联\n(金融场景)",
        "垂直场景深耕\n差异化壁垒",
        "投资决策\n量化选股",
        "付费能力强\n蓝海市场",
        "规划中",
        "0%",
        "企业版API"
    ),
]

# ─── 写表头 ──────────────────────────────────────────
for col_idx, header in enumerate(headers, start=1):
    make_header_cell(ws, 1, col_idx, header)

# ─── 写数据 ──────────────────────────────────────────
status_colors = {
    "上线":       COL_STATUS_DONE,
    "dev可运行":  COL_STATUS_DEV,
    "正在开发":   "3498DB",      # 蓝色-进行中
    "规划中":     "95A5A6",      # 灰色-规划
}

completion_colors = {
    "95%": "27AE60",
    "90%": "2ECC71",
    "85%": "58D68D",
    "80%": "82E0AA",
    "75%": "ABEBC6",
    "60%": "F9E79F",
    "20%": "FAD7A0",
    "15%": "F5B041",
    "10%": "F8C471",
    "5%":  "FDEBD0",
    "0%":  "E5E7E9",
}

for row_idx, row_data in enumerate(data, start=2):
    for col_idx, value in enumerate(row_data, start=1):
        if col_idx == 5:  # 开发程度 - 特殊样式
            cell = make_data_cell(ws, row_idx, col_idx, value, align="center")
            # 根据状态设置背景色
            for status, color in status_colors.items():
                if status in value:
                    cell.fill = make_fill(color)
                    cell.font = Font(name="Microsoft YaHei", bold=True, color="FFFFFF", size=10)
                    break
        elif col_idx == 6:  # 完成度 - 特殊样式
            cell = make_data_cell(ws, row_idx, col_idx, value, align="center")
            for pct, color in completion_colors.items():
                if pct in value:
                    cell.fill = make_fill(color)
                    cell.font = Font(name="Microsoft YaHei", bold=True, size=10)
                    break
        elif col_idx == 7:  # 商业模式
            cell = make_data_cell(ws, row_idx, col_idx, value, align="center", bold=True)
            if "¥" in value:
                cell.font = Font(name="Microsoft YaHei", bold=True, color="C0392B", size=10)
        else:
            make_data_cell(ws, row_idx, col_idx, value)

# ─── 列宽设置 ────────────────────────────────────────
col_widths = {
    1: 22,   # 功能模块
    2: 18,   # 产品定位
    3: 22,   # 应用场景
    4: 20,   # 市场需求
    5: 14,   # 开发程度
    6: 12,   # 完成度
    7: 22,   # 商业模式
}
for col, width in col_widths.items():
    ws.column_dimensions[get_column_letter(col)].width = width

# ─── 行高设置 ────────────────────────────────────────
ws.row_dimensions[1].height = 32  # header
for r in range(2, len(data) + 2):
    ws.row_dimensions[r].height = 42

# ─── 合并单元格（第一列功能模块，按阶段分组）──────────
# 给表格加标题
ws.insert_rows(1)
title_cell = ws.cell(row=1, column=1, value="SpideHarness Agent — 产品矩阵总览")
title_cell.font = Font(name="Microsoft YaHei", bold=True, size=14, color="1E3A5F")
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
ws.row_dimensions[1].height = 36

# ─── 冻结首行 ────────────────────────────────────────
ws.freeze_panes = "A3"

# ─── 添加图例 sheet ─────────────────────────────────
ws2 = wb.create_sheet("状态图例")
legend_data = [
    ("开发程度", ""),
    ("上线", "功能稳定，可投入使用", COL_STATUS_DONE),
    ("dev可运行", "功能可用，细节待优化", COL_STATUS_DEV),
    ("正在开发", "功能开发中，预计近期完成", "3498DB"),
    ("规划中", "未来规划，暂未启动", "95A5A6"),
    ("", ""),
    ("完成度", ""),
    ("90-100%", "接近完成/完成", "27AE60"),
    ("75-89%", "功能基本完成", "82E0AA"),
    ("50-74%", "功能过半", "F9E79F"),
    ("20-49%", "功能初具雏形", "FAD7A0"),
    ("<20%", "刚启动/早期阶段", "FDEBD0"),
    ("0%", "规划中，未启动", "E5E7E9"),
]
for r, (label, desc, *color) in enumerate(legend_data, start=1):
    c1 = ws2.cell(row=r, column=1, value=label)
    c2 = ws2.cell(row=r, column=2, value=desc)
    c1.font = Font(name="Microsoft YaHei", bold=bool(color))
    c2.font = Font(name="Microsoft YaHei")
    if color:
        c1.fill = make_fill(color[0])
        c1.font = Font(name="Microsoft YaHei", bold=True, color="FFFFFF")

ws2.column_dimensions["A"].width = 16
ws2.column_dimensions["B"].width = 40

# ─── 保存 ────────────────────────────────────────────
output_path = "/Users/mac/cz_code/a_Spide_agent/docs/SpideHarness产品矩阵.xlsx"
wb.save(output_path)
print(f"✅ 已生成: {output_path}")

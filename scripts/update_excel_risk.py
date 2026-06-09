#!/usr/bin/env python3
"""为产品矩阵 Excel 增加风险评估表"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook("/Users/mac/cz_code/a_Spide_agent/docs/SpideHarness产品矩阵.xlsx")

# ─── 风险评估 Sheet ─────────────────────────────────
ws_risk = wb.create_sheet("风险评估", 1)  # 插到第二页

# 配色方案（5 级风险：红→橙→黄→蓝→绿）
RISK_LEVELS = [
    {
        "level": "P0",
        "name": "极高风险",
        "color": "C0392B",
        "fg": "FFFFFF",
        "desc": "可能导致项目失败或严重损失，必须立即处理",
        "examples": "核心数据源 API 不可用 / 重大架构缺陷 / 关键依赖停止维护",
    },
    {
        "level": "P1",
        "name": "高风险",
        "color": "E67E22",
        "fg": "FFFFFF",
        "desc": "对项目目标有重大影响，需优先处理",
        "examples": "竞品大幅降价 / 技术选型失误导致重构 / 核心模块 BUG",
    },
    {
        "level": "P2",
        "name": "中风险",
        "color": "F1C40F",
        "fg": "000000",
        "desc": "对进度或质量有影响，需纳入计划",
        "examples": "第三方服务不稳定 / 部分功能实现复杂度超预期",
    },
    {
        "level": "P3",
        "name": "低风险",
        "color": "3498DB",
        "fg": "FFFFFF",
        "desc": "影响有限，可容忍或忽略",
        "examples": "UI 细节优化 / 边缘用例处理 / 文档完善",
    },
    {
        "level": "P4",
        "name": "可忽略",
        "color": "27AE60",
        "fg": "FFFFFF",
        "desc": "几乎无影响，仅需观察",
        "examples": "日志格式调整 / 代码注释优化 / 非关键依赖更新",
    },
]

# 项目实际风险
PROJECT_RISKS = [
    # (风险项, 类别, 级别, 影响模块, 当前状态, 应对策略)
    (
        "UApiPro API 稳定性",
        "外部依赖",
        "P1",
        "热搜采集 / 深度采集",
        "UAPI 免费版有调用频率限制（30 RPM），高并发场景受限",
        "升级付费版 / 接入多数据源备份 / 降级兜底方案"
    ),
    (
        "MediaCrawler 平台适配",
        "技术风险",
        "P1",
        "深度采集（7 平台）",
        "各平台频繁改版，Playwright 选择器易失效，维护成本高",
        "Autofix Skill 自动修复 / 专人负责平台适配更新"
    ),
    (
        "竞品功能碾压",
        "市场风险",
        "P2",
        "整体产品",
        "大厂（字节/腾讯）如推同类工具，可能快速复制核心功能",
        "差异化：AI 分析深度 + MCP 生态 + 开发者友好"
    ),
    (
        "GLM 模型成本与可用性",
        "技术风险",
        "P2",
        "AI 分析层",
        "智谱 AI API 调用成本 / 响应稳定性 / 模型更新影响",
        "模型降级兜底 / 结果缓存 / 成本监控告警"
    ),
    (
        "MQTT 自动重连未实现",
        "技术风险",
        "P2",
        "MQTT 通讯层",
        "网络波动时连接中断不会自动恢复，服务中断",
        "Phase 2 迭代中实现自动重连 + 心跳保活机制"
    ),
    (
        "Redis 连接容错",
        "技术风险",
        "P3",
        "存储层 / 缓存层",
        "Redis 服务未启动时程序直接崩溃，影响可用性",
        "优雅降级：Redis 不可用时切换到纯 SQLite 模式"
    ),
    (
        "爬虫合规法律风险",
        "合规风险",
        "P2",
        "深度采集模块",
        "robots.txt 遵守 / 用户协议合规 / 数据使用边界",
        "仅采集公开数据 + 用户协议提醒 + 法律顾问咨询"
    ),
    (
        "版本奇偶规则认知成本",
        "管理风险",
        "P4",
        "版本发布",
        "用户对 DEV 版（奇数）/ 正式版（偶数）规则不熟悉，误用 DEV 版",
        "CLI 启动提示 + 文档强调 + 发布日志明确标注"
    ),
    (
        "LLM 流式输出兼容性",
        "技术风险",
        "P3",
        "LLM 客户端",
        "chat_stream 同步迭代器与 async 事件循环兼容性问题",
        "asyncio.to_thread 隔离 / 重构为真正异步迭代器"
    ),
    (
        "Gateway 网关层空缺",
        "架构风险",
        "P3",
        "核心框架",
        "gateway/ 目录存在但为空，架构不完整",
        "Phase 1 前补齐，作为外部系统统一入口"
    ),
]

# ─── 样式工厂 ───────────────────────────────────────
def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def make_border(color="B8CCE4"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def make_header(ws, row, col, value, color="1E3A5F", fg="FFFFFF"):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name="Microsoft YaHei", bold=True, color=fg, size=11)
    cell.fill = make_fill(color)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = make_border("FFFFFF")
    return cell

def make_cell(ws, row, col, value, bold=False, align="left", bg=None, fg="000000", size=10):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name="Microsoft YaHei", bold=bold, color=fg, size=size)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border = make_border()
    if bg:
        cell.fill = make_fill(bg)
    return cell

# ─── 1. 风险等级图例 ────────────────────────────────
ws_risk.cell(row=1, column=1, value="SpideHarness Agent — 风险评估矩阵")
ws_risk.cell(row=1, column=1).font = Font(name="Microsoft YaHei", bold=True, size=14, color="1E3A5F")
ws_risk.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
ws_risk.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
ws_risk.row_dimensions[1].height = 36

# 表头
headers = ["风险等级", "名称", "说明", "典型示例", ""]
for c, h in enumerate(headers[:4], 1):
    make_header(ws_risk, 2, c, h)

# 数据
for i, risk in enumerate(RISK_LEVELS):
    r = i + 3
    bg = risk["color"]
    fg = risk["fg"]
    make_cell(ws_risk, r, 1, risk["level"], bold=True, align="center", bg=bg, fg=fg, size=12)
    make_cell(ws_risk, r, 2, risk["name"], bold=True, align="center", bg=bg, fg=fg, size=12)
    make_cell(ws_risk, r, 3, risk["desc"], bg=bg, fg=fg)
    make_cell(ws_risk, r, 4, risk["examples"], bg=bg, fg=fg)
    ws_risk.row_dimensions[r].height = 40

# ─── 2. 项目风险矩阵 ────────────────────────────────
start_row = len(RISK_LEVELS) + 5

# 标题
title_cell = ws_risk.cell(row=start_row, column=1, value="项目风险清单 — 按严重程度排序")
title_cell.font = Font(name="Microsoft YaHei", bold=True, size=12, color="1E3A5F")
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws_risk.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=6)
ws_risk.row_dimensions[start_row].height = 30

# 表头
risk_headers = ["风险项", "类别", "级别", "影响模块", "当前状态", "应对策略"]
for c, h in enumerate(risk_headers, 1):
    make_header(ws_risk, start_row + 1, c, h)

# 级别→颜色映射
level_colors = {r["level"]: (r["color"], r["fg"]) for r in RISK_LEVELS}

for i, risk in enumerate(PROJECT_RISKS):
    r = start_row + 2 + i
    item, category, level, modules, status, strategy = risk
    level_bg, level_fg = level_colors.get(level, ("FFFFFF", "000000"))

    make_cell(ws_risk, r, 1, item, bold=True)
    make_cell(ws_risk, r, 2, category, align="center")
    cell_lvl = make_cell(ws_risk, r, 3, level, bold=True, align="center", bg=level_bg, fg=level_fg, size=12)
    make_cell(ws_risk, r, 4, modules)
    make_cell(ws_risk, r, 5, status)
    make_cell(ws_risk, r, 6, strategy)
    ws_risk.row_dimensions[r].height = 48

# ─── 3. 列宽设置 ────────────────────────────────────
col_widths = [22, 12, 10, 22, 38, 36]
for c, w in enumerate(col_widths, 1):
    ws_risk.column_dimensions[get_column_letter(c)].width = w

# 冻结首行
ws_risk.freeze_panes = "A3"

# ─── 保存 ────────────────────────────────────────────
output = "/Users/mac/cz_code/a_Spide_agent/docs/SpideHarness产品矩阵.xlsx"
wb.save(output)
print(f"✅ 已更新: {output}")
